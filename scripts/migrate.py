#!/usr/bin/env python3
"""
migrate.py — one-shot v1 → v2 schema migration for MedHub (M3.5).

Reads per-resident-tab v1 workbooks from --input and writes a single
KCC_Master.xlsx + migration_report.txt to --output.

Target schema (15 columns, per MedHub v2 Architecture §3.1 / §4.1):
  A PatientName   B DOB           C MRN
  D MedName       E Dose          F Qty
  G Pharmacy      H PharmacyFax   I Doctor
  J LastFill      K DaysSupply    L NextFillDue (formula = J+K)
  M RefillsRemaining   N RefillStatus (blank)   O Verified (blank)

Source layout (per resident tab):
  Row 1: title (skipped)
  Row 2: B=PatientName, E=DOB, G or H=MRN
  Rows 3-7: refill alerts / instructions (skipped)
  Row 8: med-row column headers (skipped)
  Row 9+: med rows, one per medication; stop at first row with empty
          Medication Name (column B).

Source column → master column mapping (row 9+):
  src A (Include in Fax)    -> DROPPED (v1 transient state)
  src B (Medication Name)   -> D MedName
  src C (Dose)              -> E Dose
  src D (Quantity)          -> F Qty
  src E (Pharmacy Name)     -> G Pharmacy
  src F (Pharmacy Fax)      -> H PharmacyFax
  src G (Prescribing Doctor)-> I Doctor
  src H (Last Fill Date)    -> J LastFill
  src I (Days Supply)       -> K DaysSupply
  src J (Next Fill Due)     -> DROPPED (column L is a formula in v2)
  src K (Refills Remaining) -> M RefillsRemaining

Dependency:
  pip3 install openpyxl

Example:
  python3 ~/medhub/scripts/migrate.py \\
      --input  ~/medhub_migration/input/ \\
      --output ~/medhub_migration/output/

The migration_report.txt is PHI-free by design — counts and structural
anomalies only, no patient names, DOBs, MRNs, or med names.
"""

import argparse
import datetime as dt
import sys
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    sys.exit("openpyxl is required. Install with: pip3 install openpyxl")


MASTER_HEADERS = [
    "PatientName", "DOB", "MRN",
    "MedName", "Dose", "Qty",
    "Pharmacy", "PharmacyFax", "Doctor",
    "LastFill", "DaysSupply", "NextFillDue",
    "RefillsRemaining", "RefillStatus", "Verified",
]


def _cell(ws, coord):
    """Return a cell's value, treating blanks/whitespace-only strings as None."""
    v = ws[coord].value
    if v is None:
        return None
    if isinstance(v, str):
        v = v.strip()
        return v if v else None
    return v


def _read_patient_header(ws):
    """Return (name, dob, mrn, mrn_source) from row 2. mrn_source is 'G', 'H', or None."""
    name = _cell(ws, "B2")
    dob = _cell(ws, "E2")
    mrn_g = _cell(ws, "G2")
    mrn_h = _cell(ws, "H2")
    if mrn_g is not None:
        return name, dob, mrn_g, "G"
    if mrn_h is not None:
        return name, dob, mrn_h, "H"
    return name, dob, None, None


def _walk_med_rows(ws):
    """Yield med dicts starting at row 9, stopping at the first empty MedName (col B)."""
    row = 9
    while True:
        med_name = _cell(ws, f"B{row}")
        if med_name is None:
            return
        yield {
            "medName":          med_name,
            "dose":             _cell(ws, f"C{row}"),
            "qty":              _cell(ws, f"D{row}"),
            "pharmacy":         _cell(ws, f"E{row}"),
            "pharmacyFax":      _cell(ws, f"F{row}"),
            "doctor":           _cell(ws, f"G{row}"),
            "lastFill":         _cell(ws, f"H{row}"),
            "daysSupply":       _cell(ws, f"I{row}"),
            "refillsRemaining": _cell(ws, f"K{row}"),
        }
        row += 1


def process_workbook(path, report):
    """Parse one v1 workbook; return a list of master-row dicts."""
    rows = []
    wb = load_workbook(filename=path, data_only=True, read_only=False)
    wb_block = {
        "filename": path.name,
        "tab_count": len(wb.sheetnames),
        "residents_emitted": 0,
        "meds_emitted": 0,
        "anomalies": [],
    }

    for tab_index, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        name, dob, mrn, mrn_src = _read_patient_header(ws)

        if name is None:
            wb_block["anomalies"].append(
                f"tab {tab_index}: no patient name in B2 (tab skipped)"
            )
            continue

        meds = list(_walk_med_rows(ws))
        if not meds:
            wb_block["anomalies"].append(
                f"tab {tab_index}: patient header present, zero medication rows"
            )
            continue

        if mrn is None:
            wb_block["anomalies"].append(
                f"tab {tab_index}: no MRN in G2 or H2 (emitted with blank MRN)"
            )
        elif mrn_src == "H":
            wb_block["anomalies"].append(
                f"tab {tab_index}: MRN found in H2 instead of G2"
            )

        wb_block["residents_emitted"] += 1
        wb_block["meds_emitted"] += len(meds)

        for med in meds:
            rows.append({
                "PatientName":      name,
                "DOB":              dob,
                "MRN":              mrn,
                "MedName":          med["medName"],
                "Dose":             med["dose"],
                "Qty":              med["qty"],
                "Pharmacy":         med["pharmacy"],
                "PharmacyFax":      med["pharmacyFax"],
                "Doctor":           med["doctor"],
                "LastFill":         med["lastFill"],
                "DaysSupply":       med["daysSupply"],
                "RefillsRemaining": med["refillsRemaining"],
            })

    wb.close()
    report["workbooks"].append(wb_block)
    return rows


def _sort_key(r):
    pn = r.get("PatientName") or ""
    mn = r.get("MedName") or ""
    return (str(pn).lower(), str(mn).lower())


def write_master(path, all_rows):
    """Write KCC_Master.xlsx: title, timestamp, headers, data, named ListObject."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Medications"

    ws["A1"] = "KCC Master Medication Tracker — Kelly Cullen Community"
    ws["A2"] = f"Generated {dt.datetime.now().isoformat(timespec='seconds')} by migrate.py v1"

    for col_idx, header in enumerate(MASTER_HEADERS, start=1):
        ws.cell(row=3, column=col_idx, value=header)

    sorted_rows = sorted(all_rows, key=_sort_key)

    for i, r in enumerate(sorted_rows):
        excel_row = 4 + i
        ws.cell(row=excel_row, column=1,  value=r["PatientName"])
        ws.cell(row=excel_row, column=2,  value=r["DOB"])
        ws.cell(row=excel_row, column=3,  value=r["MRN"])
        ws.cell(row=excel_row, column=4,  value=r["MedName"])
        ws.cell(row=excel_row, column=5,  value=r["Dose"])
        ws.cell(row=excel_row, column=6,  value=r["Qty"])
        ws.cell(row=excel_row, column=7,  value=r["Pharmacy"])
        ws.cell(row=excel_row, column=8,  value=r["PharmacyFax"])
        ws.cell(row=excel_row, column=9,  value=r["Doctor"])
        ws.cell(row=excel_row, column=10, value=r["LastFill"])
        ws.cell(row=excel_row, column=11, value=r["DaysSupply"])
        ws.cell(row=excel_row, column=12,
                value=f'=IF(AND(J{excel_row}<>"",K{excel_row}<>""),J{excel_row}+K{excel_row},"")')
        ws.cell(row=excel_row, column=13, value=r["RefillsRemaining"])
        # N RefillStatus and O Verified are intentionally left blank (v2 UX additions)

    last_row = 3 + len(sorted_rows)
    if last_row >= 4:
        tbl = Table(displayName="Medications", ref=f"A3:O{last_row}")
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(tbl)

    wb.save(path)


def write_report(path, report):
    """Write PHI-free migration_report.txt (counts and structural anomalies only)."""
    totals = {
        "workbooks":         len(report["workbooks"]),
        "tabs":              sum(w["tab_count"]         for w in report["workbooks"]),
        "residents_emitted": sum(w["residents_emitted"] for w in report["workbooks"]),
        "meds_emitted":      sum(w["meds_emitted"]      for w in report["workbooks"]),
    }

    lines = []
    lines.append("KCC MedHub v1 -> v2 Migration Report")
    lines.append(f"Generated: {dt.datetime.now().isoformat(timespec='seconds')}")
    lines.append("")
    lines.append("Totals")
    lines.append("------")
    lines.append(f"Workbooks processed:   {totals['workbooks']}")
    lines.append(f"Tabs walked:           {totals['tabs']}")
    lines.append(f"Residents emitted:     {totals['residents_emitted']}")
    lines.append(f"Medication rows:       {totals['meds_emitted']}")
    lines.append("")
    lines.append("Per workbook")
    lines.append("------------")
    for wb in report["workbooks"]:
        lines.append(f"{wb['filename']}")
        lines.append(f"  tabs:              {wb['tab_count']}")
        lines.append(f"  residents emitted: {wb['residents_emitted']}")
        lines.append(f"  meds emitted:      {wb['meds_emitted']}")
        if wb["anomalies"]:
            lines.append("  anomalies:")
            for a in wb["anomalies"]:
                lines.append(f"    - {a}")
        lines.append("")

    path.write_text("\n".join(lines) + "\n")


def main():
    p = argparse.ArgumentParser(
        description="Migrate v1 per-resident-tab workbooks to KCC_Master.xlsx"
    )
    p.add_argument("--input",  required=True, type=Path, help="Directory of source .xlsx workbooks")
    p.add_argument("--output", required=True, type=Path, help="Directory to write outputs")
    p.add_argument("--dry-run", action="store_true",
                   help="Parse and report only; do not write KCC_Master.xlsx or migration_report.txt")
    p.add_argument("--verbose", action="store_true",
                   help="Print per-workbook anomaly details to stdout (PHI-free; tab indices only)")
    args = p.parse_args()

    args.input = args.input.expanduser()
    args.output = args.output.expanduser()

    if not args.input.is_dir():
        sys.exit(f"Input directory not found: {args.input}")

    sources = sorted(p for p in args.input.glob("*.xlsx") if not p.name.startswith("~$"))
    if not sources:
        sys.exit(f"No .xlsx files found in {args.input}")

    print(f"Found {len(sources)} workbook(s) in {args.input}")
    if args.dry_run:
        print("DRY RUN — no files will be written")

    report = {"workbooks": []}
    all_rows = []

    for src in sources:
        print(f"  reading {src.name} …")
        rows = process_workbook(src, report)
        all_rows.extend(rows)

    total_anomalies = sum(len(w["anomalies"]) for w in report["workbooks"])
    print()
    print(f"Totals: {len(report['workbooks'])} workbooks, "
          f"{sum(w['tab_count'] for w in report['workbooks'])} tabs, "
          f"{sum(w['residents_emitted'] for w in report['workbooks'])} residents, "
          f"{len(all_rows)} medication rows, "
          f"{total_anomalies} anomalies")

    if args.verbose:
        print()
        print("Anomalies (PHI-free — tab indices only)")
        print("----------------------------------------")
        for wb in report["workbooks"]:
            if not wb["anomalies"]:
                continue
            print(f"{wb['filename']} (tabs={wb['tab_count']}, residents={wb['residents_emitted']}, meds={wb['meds_emitted']}):")
            for a in wb["anomalies"]:
                print(f"  - {a}")

    if args.dry_run:
        print("\nDRY RUN complete. Re-run without --dry-run to write outputs.")
        return

    args.output.mkdir(parents=True, exist_ok=True)
    master_path = args.output / "KCC_Master.xlsx"
    report_path = args.output / "migration_report.txt"

    write_master(master_path, all_rows)
    write_report(report_path, report)

    print(f"\nWrote {master_path}")
    print(f"Wrote {report_path}")


if __name__ == "__main__":
    main()
